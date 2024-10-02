import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from tkinter import messagebox
from tkinter import StringVar
from datetime import datetime
import pandas as pd

# Função para formatar a data
def formatar_data(data):
    if len(data) == 6 and data.isdigit():
        return f"{data[:2]}/{data[2:4]}/{data[4:]}"
    return data

# Função para atualizar a entrada de data
def atualizar_data(event):
    entrada_data = entry_data.get()
    data_formatada = formatar_data(entrada_data)
    entry_data.delete(0, 'end')
    entry_data.insert(0, data_formatada)

# Função para calcular e salvar na planilha
def calcular_e_salvar():
    nome_produto = entry_nome.get()
    valor_compra = entry_valor_compra.get()
    valor_venda = entry_valor_venda.get()
    quantidade = entry_quantidade.get()
    observacao = entry_observacao.get("1.0", "end")
    data_entrada = entry_data.get()

    if not nome_produto or not valor_compra or not valor_venda or not quantidade or not data_entrada:
        messagebox.showerror("Erro", "Por favor, preencha todos os campos obrigatórios.")
        return

    try:
        valor_compra = int(valor_compra)
        valor_venda = int(valor_venda)
        quantidade = int(quantidade)
        # Verifica se a data está no formato correto
        data_entrada = datetime.strptime(data_entrada, '%d/%m/%y')
    except ValueError:
        messagebox.showerror("Erro", "Os valores de compra, venda, quantidade e a data devem ser válidos.")
        return

    # Cálculo de lucro e faturamento
    lucro = (valor_venda - valor_compra)
    faturamento = valor_venda

    try:
        file_name = "relatorio_produtos.xlsx"
        
        # Tenta carregar uma planilha existente
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            # Se não existir, cria um novo
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Relatório de Produtos"
            # Cabeçalhos
            headers = ["Nome do Produto", "Data da Entrada", "Quantidade", "Valor de Compra", "Valor de Venda", "Faturamento", "Lucro", "Observação", "Gastos Totais", "Faturamento Total", "Lucro Total"]
            sheet.append(headers)
        
        # Adiciona dados
        data = [nome_produto, data_entrada.strftime('%d/%m/%y'), quantidade, valor_compra, valor_venda, faturamento, lucro, observacao.strip()]
        sheet.append(data)

        # Formatação dos cabeçalhos (se novo arquivo)
        if sheet.max_row == 2:  # Se a planilha é nova (somente cabeçalhos presentes)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Ajuste das larguras das colunas
            column_widths = [max(len(str(item)) for item in column) for column in zip(*sheet.iter_rows(values_only=True))]
            for i, column_width in enumerate(column_widths, start=1):
                column_letter = chr(64+i)
                sheet.column_dimensions[column_letter].width = column_width + 2

        workbook.save(file_name)
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso na planilha relatorio_produtos.xlsx")
        
        # Limpar todos os campos
        entry_nome.delete(0, 'end')
        entry_quantidade.delete(0, 'end')
        entry_valor_compra.delete(0, 'end')
        entry_valor_venda.delete(0, 'end')
        entry_observacao.delete("1.0", 'end')
        entry_data.delete(0, 'end')
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao salvar a planilha: {str(e)}")

# Função para calcular e exibir os totais mensais
def calcular_totais_mensais():
    file_name = "relatorio_produtos.xlsx"
    try:
        # Carregar os dados da planilha
        df = pd.read_excel(file_name, sheet_name="Relatório de Produtos")

        # Exibir dados carregados para depuração
        print("Dados Carregados: ")
        print(df.head())

        # Converter a coluna de data para o formato datetime
        df['Data da Entrada'] = pd.to_datetime(df['Data da Entrada'], format='%d/%m/%y')

        # Identificar o primeiro e o último dia das vendas do mês
        data_minima = df ['Data da Entrada'].min()
        data_maxima = df ['Data da Entrada'].max()

        # Exibir data minima e maxima para a depuração
        print(f"Data Minima: {data_minima}")
        print(f"Data Maxima: {data_maxima}")

        # Filtrar dados do primeiro ao último dia do mês atual
        inicio_mes = data_minima.replace(day=1)
        fim_mes = (data_minima + pd.DateOffset(months=1)).replace(day=1) - pd.DateOffset(days=1)

        # Filtrar os dados
        df_filtrado = df[(df['Data da Entrada'] >= inicio_mes) & (df['Data da Entrada'] <= fim_mes)]

        # Calcular o total de gastos, lucros e faturamentos
        totais = df_filtrado[['Lucro', 'Faturamento']].sum()
        lucro_total = totais['Lucro']
        faturamento_total = totais['Faturamento']
        gastos_totais = df_filtrado['Gastos Totais'].sum()

        # Atualizar a interface com os resultados
        label_resultado.configure(text=f"Lucro Total: R${lucro_total:.2f}\nFaturamento Total: R${faturamento_total:.2f}")

        # Adicionar ou atualizar a linha com totais mensais na planilha
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active

            # Verificar se a linha de totais já existe
            found_totals_row = False
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                if row[0].value == "Totais Mensais":
                    found_totals_row = True
                    break

            if not found_totals_row:
                # Adicionar nova linha com totais mensais
                sheet.append([
                    "Totais Mensais", "", "", "", "", "", "", "",
                    gastos_totais, faturamento_total, lucro_total
                ])
            else:
                # Atualizar a linha existente com totais mensais
                totals_row = sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)
                for row in totals_row:
                    if row[0].value == "Totais Mensais":
                        row_index = row[0].row
                        sheet.cell(row=row_index, column=10, value=faturamento_total)  # Faturamento Total
                        sheet.cell(row=row_index, column=11, value=lucro_total)  # Lucro Total
                        break

            # Formatar a linha de totais
            for cell in sheet[sheet.max_row]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            workbook.save(file_name)
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao adicionar totais na planilha: {str(e)}")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao calcular os totais mensais: {str(e)}")

            
# Função para alternar o modo de aparência
def alternar_tema():
    global tema_atual
    if tema_atual == "dark":
        ctk.set_appearance_mode("light")
        btn_alternar_tema.configure(text="Modo Escuro")
        tema_atual = "light"
    else:
        ctk.set_appearance_mode("dark")
        btn_alternar_tema.configure(text="Modo Claro")
        tema_atual = "dark"

# Configurando a interface gráfica com CustomTkinter
ctk.set_appearance_mode("dark")  # Modo escuro inicial
ctk.set_default_color_theme("green")  # Tema padrão

tema_atual = "dark"

root = ctk.CTk()  # Janela principal
root.title("Gestão de Produtos")
root.geometry("800x600")  # Tamanho fixo da janela
root.resizable(False, False)
root.grid_rowconfigure(0, weight=1)  # Centraliza verticalmente
root.grid_columnconfigure(0, weight=1)  # Centraliza horizontalmente

# Frame principal para centralização
frame = ctk.CTkFrame(root)
frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

# Configurando os pesos das linhas e colunas para expandir e preencher
frame.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9), weight=1)
frame.grid_columnconfigure((0, 1), weight=1)

# Nome do produto
ctk.CTkLabel(frame, text="Nome do Produto:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
entry_nome = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o nome")
entry_nome.grid(row=0, column=1, padx=10, pady=10, sticky="w")

# Data da Entrada
ctk.CTkLabel(frame, text="Data da Venda:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
entry_data = ctk.CTkEntry(frame, width=200, placeholder_text="Digite a data (DDMMYY)")
entry_data.grid(row=1, column=1, padx=10, pady=10, sticky="w")
entry_data.bind("<FocusOut>", atualizar_data)  # Evento de perda de foco para formatar a data

# Quantidade
ctk.CTkLabel(frame, text="Quantidade:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
entry_quantidade = ctk.CTkEntry(frame, width=200, placeholder_text="Digite a quantidade")
entry_quantidade.grid(row=2, column=1, padx=10, pady=10, sticky="w")

# Valor de Compra
ctk.CTkLabel(frame, text="Valor de Compra:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
entry_valor_compra = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o valor de compra")
entry_valor_compra.grid(row=3, column=1, padx=10, pady=10, sticky="w")

# Valor de Venda
ctk.CTkLabel(frame, text="Valor de Venda:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
entry_valor_venda = ctk.CTkEntry(frame, width=200, placeholder_text="Digite o valor de venda")
entry_valor_venda.grid(row=4, column=1, padx=10, pady=10, sticky="w")

# Observação
ctk.CTkLabel(frame, text="Observação:").grid(row=5, column=0, padx=10, pady=10, sticky="ne")
entry_observacao = ctk.CTkTextbox(frame, width=200, height=100)
entry_observacao.grid(row=5, column=1, padx=10, pady=10, sticky="w")

# Botão Calcular
btn_calcular = ctk.CTkButton(frame, text="Calcular e Salvar", command=calcular_e_salvar)
btn_calcular.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

# Botão Calcular Totais
btn_calcular_totais = ctk.CTkButton(frame, text="Calcular Totais Mensais", command=calcular_totais_mensais)
btn_calcular_totais.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

# Resultado
label_resultado = ctk.CTkLabel(frame, text="Lucro Total: R$0.00\nFaturamento Total: R$0.00")
label_resultado.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

# Botão Alternar Tema
btn_alternar_tema = ctk.CTkButton(frame, text="Modo Claro", command=alternar_tema)
btn_alternar_tema.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()